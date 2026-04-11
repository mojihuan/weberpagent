# coding: utf-8
import json
import os
import time
import inspect
from functools import wraps
import re
import pyperclip
from bs4 import BeautifulSoup

from common.base_assert import BaseAssert
from common.base_params import InitializeParams
from common.base_smart_positioning import BaseSmartPositioning, desc_collector, COMPILED_PATTERNS
from common.base_url import URL
from openpyxl import load_workbook
from selenium.common import ElementNotVisibleException, \
    InvalidElementStateException, InvalidSelectorException, ElementClickInterceptedException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta
from common.element_positioning import ElementPositioning
from common.base_random_mixin import BaseRandomMixin
from common.import_api import ImportApi
from config.settings import DATA_PATHS
from selenium.common import TimeoutException, NoSuchElementException

_global_step_counter = 1  # 初始化递增步骤


class BasePage(BaseRandomMixin, BaseSmartPositioning):
    """
    页面基础类，封装了 Selenium WebDriver 的等常用操作方法
    """

    def __init__(self, driver, *args, **kwargs):
        """
        初始化方法，设置登录页面 URL、驱动实例、动作链等
        :param driver: WebDriver 实例
        """
        super().__init__(*args, **kwargs)
        self.login_url_loc = URL['web']['login'].replace('/epbox_erp', '')  # ui 登录 url
        self.driver = driver  # 谷歌驱动实例
        self.actions = ActionChains(self.driver)  # 键盘事件
        self.base_assert = BaseAssert()  # 公共断言类
        self.file_paths = {}  # 文件路径
        self.elem_positioning = ElementPositioning.__dict__  # 公共元素定位
        self.positioning = self.elem_positioning['positioning']  # 元素定位字典（简化访问）
        self.api = ImportApi()  # api 查询类
        self.URL = URL['api']  # 公共请求地址
        self._captured_requests_map = {}  # 使用字典可实现相同 keyword 自动覆盖，确保只保留最后一次请求

    def get_formatted_datetime(self, days=0, hours=0):
        """
        获取当前时间，并支持增加指定天数和小时数
        :param days: 要增加的天数（默认为 0）
        :param hours: 要增加的小时数（默认为 0）
        :return: 格式化后的时间字符串
        """
        current_time = datetime.now() + timedelta(days=days, hours=hours)
        return current_time.strftime("%Y-%m-%d %H:%M:%S")

    def get_current_time_hms(self):
        """
        获取当前时分秒
        :return: 格式化后的时间字符串（时分秒）
        """
        return time.strftime("%H:%M:%S", time.localtime(time.time()))

    def get_the_date(self, days=0):
        """
        获取当前日期，并支持增加指定天数
        :param days: 要增加的天数（默认为 0）
        :return: 格式化后的日期字符串（仅年月日）
        """
        current_time = datetime.now() + timedelta(days=days)
        return current_time.strftime("%Y-%m-%d")

    def get_load(self):
        """
        打开登录页面
        :return: 当前对象自身
        """
        self.driver.get(self.login_url_loc)
        return self

    def get_time_stamp(self):
        """
        获取当前时间戳（毫秒级）
        :return: 时间戳字符串
        """
        ct = time.time()
        local_time = time.localtime(ct)
        data_head = time.strftime("%Y%m%d%H%M%S", local_time)
        data_secs = (ct - int(ct)) * 1000
        time_stamp = "%s%03d" % (data_head, data_secs)
        return time_stamp

    def wait_time(self, seconds=0.8):
        """
        强制等待指定时间
        :param seconds: 等待秒数
        :return: 当前对象自身
        """
        time.sleep(seconds)
        return self

    def move(self, key=None):
        """
        等待元素可见并定位
        :param key: 元素定位器 (By, value) 或定位器列表
        :return: 定位到的元素
        """

        def find_with_exc():
            selectors = key if isinstance(key, list) else [key]

            for sel in selectors:
                locator = self._identify_locator(sel)
                wait = WebDriverWait(self.driver, DATA_PATHS['wait_time'], poll_frequency=0.5)
                try:
                    return wait.until(EC.visibility_of_element_located(locator))
                except (TimeoutException, NoSuchElementException):
                    continue

            raise NoSuchElementException(f"无法找到任何元素：{selectors}")

        return self.exc(find_with_exc)

    def css(self, key=None):
        """
        使用 CSS 选择器定位元素
        :param key: CSS 选择器字符串
        :return: By.CSS_SELECTOR 和选择器值
        """

        def find_with_exc():
            return By.CSS_SELECTOR, key

        return self.exc(find_with_exc)

    def xp(self, key=None):
        """
        使用 XPath 选择器定位元素
        :param key: XPath 字符串
        :return: By.XPATH 和选择器值
        """

        def find_with_exc():
            return By.XPATH, key

        return self.exc(find_with_exc)

    def find_elem(self, key=None):
        """
        查找元素，直接定位，不等待其存在于 DOM 中
        :param key: 元素定位器 (By, value) 或定位器列表
        :return: 定位到的元素
        """

        def find_with_exc():
            selectors = key if isinstance(key, list) else [key]
            for sel in selectors:
                locator = self._identify_locator(sel)
                try:
                    return self.driver.find_element(*locator)
                except (NoSuchElementException, InvalidSelectorException):
                    continue
            raise NoSuchElementException(f"无法找到任何元素：{selectors}")

        return self.exc(find_with_exc)

    def exc(self, func):
        """
        异常处理方法
        :param func: 待执行的函数
        :return: 函数执行结果或抛出异常
        """
        exception_handlers = {
            TimeoutException: "【元素加载超时】：元素未在指定时间内加载完成，请增加等待时间或检查网络",
            NoSuchElementException: "【元素未找到】：未在页面上找到目标元素，请检查定位器是否正确或页面是否已加载完成",
            InvalidSelectorException: "【定位器表达式错误】：XPath/CSS 表达式不合法，请检查语法是否正确",
            ElementNotVisibleException: "【元素不可见】：元素存在于 DOM 中，但不可见 可能被隐藏",
            InvalidElementStateException: "【元素状态异常】：元素处于只读、禁用或不可交互状态，无法执行操作",
            ElementClickInterceptedException: "【元素点击被拦截】：该元素可能被弹窗或其他组件遮挡，请检查是否有弹窗、对话框等影响点击",
            KeyError: "【元素 key 不存在】：定位器配置中找不到指定的 key，请确认元素 key 是否拼写错误或未定义"
        }
        try:
            return func()
        except Exception as e:
            msg = exception_handlers.get(type(e), "未知异常")
            if hasattr(e, 'msg'):
                detail = e.msg
            elif hasattr(e, 'args') and len(e.args) > 0:
                detail = str(e.args[0])
            else:
                detail = str(e)
            print(f"{msg}: {detail}")
            self.take_step_screenshot("error")
            raise  # 抛出原始异常，不影响后续行为

    def up_enter(self, count=1):
        """
        键盘 - 向上箭头 - 回车
        :param count: 按向上箭头的次数
        :return: None
        """
        self.wait_time()
        for _ in range(count):
            self.actions.send_keys(Keys.ARROW_UP).perform()
        self.actions.key_down(Keys.RETURN).key_up(Keys.RETURN).perform()

    def down_enter(self, count=1):
        """
        键盘 - 向下箭头 - 回车
        :param count: 按向下箭头的次数
        :return: None
        """
        self.wait_time()
        for _ in range(count):
            self.actions.send_keys(Keys.ARROW_DOWN).perform()
        self.actions.key_down(Keys.RETURN).key_up(Keys.RETURN).perform()

    def down_right_right_enter(self):
        """
        键盘 - 向下箭头 - 回车
        :return: None
        """
        self.wait_time()
        self.actions.send_keys(Keys.ARROW_DOWN).perform()
        self.actions.send_keys(Keys.ARROW_RIGHT).perform()
        self.actions.send_keys(Keys.ARROW_RIGHT).perform()
        self.actions.key_down(Keys.RETURN).key_up(Keys.RETURN).perform()

    def tab_space(self, count=1):
        """
        键盘-Tab 键 - 空格键
        :param count: 按 Tab 键的次数
        :return: None
        """
        self.wait_time()
        for _ in range(count):
            self.actions.send_keys(Keys.TAB).perform()
        self.actions.send_keys(Keys.SPACE).perform()

    def tab_enter(self, count=1):
        """
        键盘-Tab 键 - 回车键
        :param count: 按 Tab 键的次数
        :return: None
        """
        self.wait_time()
        for _ in range(count):
            self.actions.send_keys(Keys.TAB).perform()
        self.actions.key_down(Keys.RETURN).key_up(Keys.RETURN).perform()

    def tab(self, count=1):
        """
        键盘-Tab 键
        :param count: 按 Tab 键的次数
        :return: None
        """
        self.wait_time()
        for _ in range(count):
            self.actions.send_keys(Keys.TAB).perform()

    def enter(self):
        """
        键盘 - 回车
        :return: None
        """
        self.wait_time()
        self.actions.key_down(Keys.RETURN).key_up(Keys.RETURN).perform()

    def esc(self):
        """
        键盘 - 按 ESC 键
        :return: None
        """
        self.wait_time()
        self.actions.key_down(Keys.ESCAPE).key_up(Keys.ESCAPE).perform()

    def ctrl_v(self):
        """
        键盘 - 粘贴
        :return: None
        """
        self.wait_time()
        self.actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()

    def ctrl_v_enter(self):
        """
        键盘 - 粘贴 - 回车
        :return: None
        """
        self.wait_time()
        self.actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
        self.actions.key_down(Keys.RETURN).key_up(Keys.RETURN).perform()

    def take_screenshot(self, step_num, step_name):
        """
        截图并保存到指定目录
        :param step_num: 步骤编号
        :param step_name: 步骤名称
        :return: None
        """
        # 检查截图功能是否开启
        if not DATA_PATHS.get('screenshot_enabled', True):
            return

        try:
            test_caller_frame = None
            # 遍历调用栈，找到以 test_ 开头的测试方法
            for frame in inspect.stack()[2:]:
                if frame.function.startswith("test_"):
                    test_caller_frame = frame
                    break

            if not test_caller_frame:
                # 如果不是测试方法调用，使用当前调用者信息
                caller_frame = inspect.stack()[2]
                caller_file_path = caller_frame.filename
                caller_method_name = caller_frame.function
            else:
                # 获取调用者文件路径和方法名
                caller_file_path = test_caller_frame.filename
                caller_method_name = test_caller_frame.function

            caller_file_name = os.path.splitext(os.path.basename(caller_file_path))[0]

            # 构建截图目录路径（基于文件名和方法名）
            screenshot_dir = os.path.join("img", caller_file_name, caller_method_name)
            if not os.path.exists(screenshot_dir):
                os.makedirs(screenshot_dir)

            # 清理步骤名称中的非法字符
            safe_step_name = step_name.replace("/", "_").replace("\\", "_").replace(":", "_")
            filename = os.path.join(screenshot_dir, f"step_{step_num}__{safe_step_name}.png")

            # 检查 driver 是否存在
            if not hasattr(self, 'driver') or self.driver is None:
                return

            # 保存截图 - 使用绝对路径
            abs_filename = os.path.abspath(filename)

            try:
                result = self.driver.save_screenshot(abs_filename)

                # 检查文件是否真的存在
                if os.path.exists(abs_filename):
                    file_size = os.path.getsize(abs_filename)
                else:
                    raise Exception(f"截图文件未创建：{abs_filename}")

            except Exception as save_error:
                # 尝试使用 get_screenshot_as_png 方式
                try:
                    with open(abs_filename, 'wb') as f:
                        f.write(self.driver.get_screenshot_as_png())
                except Exception as backup_error:
                    raise Exception(f"备用方案也失败：{str(backup_error)}")

        except Exception as e:
            raise

    def take_step_screenshot(self, description):
        """
        自动递增步骤编号并截图
        :param description: 截图描述
        :return: 步骤编号
        """
        global _global_step_counter
        step_num = _global_step_counter
        self.take_screenshot(step_num, description)
        _global_step_counter += 1
        return step_num

    def refresh_page(self):
        """
        刷新当前页面
        :return: None
        """
        current_url = self.driver.current_url
        self.driver.get(current_url)

    def driver_back(self):
        """
        浏览器后退功能
        :return: None
        """
        self.driver.back()

    def copy(self, value):
        """
        将指定值复制到剪贴板
        :param value: 要复制到剪贴板的值
        :return: 返回自身实例以支持链式调用
        """
        pyperclip.copy(str(value))
        return self

    def click(self, key=None, force_js=False, desc=None, auto=None, index=1, tag=None, p_tag=None, p_name=None, o_tag=None, o_name=None, d=None):
        """
       单击指定元素
        :param key: 元素定位器 (By, value) 或定位器列表，或元素 key 字符串
        :param force_js: 是否强制使用 JavaScript 点击
        :param desc: 注解以及截图的名称（可选，未提供时自动生成）
        :param auto: 是否自动生成并更新定位（默认从 settings.py 读取，也可手动覆盖）
        :param index: 元素索引（从 1 开始，用于点击第 n 个重复元素）
        :param tag: HTML 标签名（如 'button', 'input', 'span' 等），用于与 desc 组合匹配
        :param p_tag: 父元素标签类型（可选）
        :param p_name: 父元素类名（可选）
        :param o_tag: 祖父元素标签类型（可选）
        :param o_name: 祖父元素属性值（如 aria-label）
        :param d: 额外的注释，没代码逻辑
        :return: None
        """
        self.wait_time()

        # 如果 auto 参数未提供，则从配置文件读取默认值
        if auto is None:
            auto = DATA_PATHS.get('auto_generate_xpath', True)

        def find_with_exc():
            # 如果传入的是字符串且不是定位器元组，则从 positioning 中获取
            if isinstance(key, str) and not key.startswith(('//', '/', 'xpath=', '(')):
                selector_value = self.positioning.get(key)
                if selector_value is None:
                    raise KeyError(f"元素 key '{key}' 不存在")
                key_name = key  # 记录使用的 key 名称
            else:
                selectors = key if isinstance(key, list) else [key]
                selector_value = selectors
                key_name = None

            # 统一处理为列表
            selectors_to_try = selector_value if isinstance(selector_value, list) else [selector_value]

            wait = WebDriverWait(self.driver, 5, poll_frequency=0.5)

            for selector_item in selectors_to_try:
                # 跳过空字符串
                if not selector_item or (isinstance(selector_item, str) and not selector_item.strip()):
                    continue

                locator = self._identify_locator(selector_item)
                try:
                    # 如果指定了 index，使用 XPath 索引
                    if index is not None and isinstance(locator, tuple) and locator[0] == By.XPATH:
                        # 检查 XPath 是否已经包含索引
                        if not re.search(r'\[\d+\]$', locator[1]):
                            # 添加索引（使用括号格式）
                            locator = (By.XPATH, f"({locator[1]})[{index}]")
                            # print(f"🎯 使用 XPath 索引：({locator[1]})[{index}]")

                    element = wait.until(EC.element_to_be_clickable(locator))
                    self.scroll(element=element)

                    if not force_js:
                        element.click()
                    else:
                        self.driver.execute_script("arguments[0].click();", element)

                    # 截图
                    screenshot_desc = desc if desc else f"点击：{key_name if key_name else locator}"
                    if index:
                        screenshot_desc = f"点击：{key_name if key_name else locator} (第{index}个)"
                    self.take_step_screenshot(screenshot_desc)

                    # 收集 desc 信息用于后续生成 XPath
                    if desc and key_name and auto:
                        # 先检查是否已有 HTML 文件路径
                        html_path = None
                        if desc_collector:
                            last_record = desc_collector[-1]
                            if last_record.get('html_path') and os.path.exists(last_record['html_path']):
                                html_path = last_record['html_path']

                        # 如果没有有效的 HTML 文件，保存一个新的（首次获取）
                        if not html_path:
                            # print(f"📄 首次保存 HTML - URL: {self.driver.current_url}")
                            html_path = self.save_html_code(file_name='elem/auto_generated.html')

                        # 收集 desc 信息（带上 HTML 路径、index、tag 和父元素、祖父元素信息）
                        self.collect_desc_info(key_name, desc, html_path, index=index, tag=tag, p_tag=p_tag, p_name=p_name, o_tag=o_tag, o_name=o_name)

                        # 自动生成并更新 XPath（实时保存，传递 index、tag 和父元素、祖父元素信息）
                        self.auto_and_update_xpath_realtime(key=key_name, desc=desc, html_path=html_path, index=index, tag=tag, p_tag=p_tag, p_name=p_name, o_tag=o_tag, o_name=o_name)
                    elif desc and key_name and not auto:
                        # 不生成定位，但记录 desc 信息
                        pass
                        # print(f"ℹ️  已跳过自动生成 - key: {key_name}, desc: {desc}")

                    # 点击成功后刷新 HTML 源代码
                    self.refresh_html_source()

                    return

                except (TimeoutException, ElementNotVisibleException, InvalidElementStateException, InvalidSelectorException) as e:
                    # 如果是非法选择器错误，打印详细信息
                    if isinstance(e, InvalidSelectorException):
                        print(f"⚠️  选择器错误 - selector_item: {selector_item}, locator: {locator}")
                    continue

            error_msg = f"无法点击所有提供的元素：{selectors_to_try}"
            raise RuntimeError(error_msg)

        return self.exc(find_with_exc)

    def input(self, key=None, text=None, desc=None, auto=None, index=1, tag=None, p_tag=None, p_name=None, o_tag=None, o_name=None, d=None):
        """
        在指定元素中输入文本，并等待其可点击
        :param key: 元素定位器 (By, value) 或定位器列表，或元素 key 字符串
        :param text: 输入的文本
        :param desc: 注解以及截图的名称（可选，未提供时自动生成）
        :param auto: 是否自动生成并更新定位（默认从 settings.py 读取，也可手动覆盖）
        :param index: 元素索引（从 1 开始，用于在第 n 个重复元素中输入）
        :param tag: HTML 标签名（如 'input', 'textarea' 等），用于与 desc 组合匹配
        :param p_tag: 父元素标签类型（可选）
        :param p_name: 父元素类名（可选）
        :param o_tag: 祖父元素标签类型（可选）
        :param o_name: 祖父元素属性值（如 aria-label）
        :param d: 额外的注释，没代码逻辑
        :return: None
        """
        self.wait_time()

        # 如果 auto 参数未提供，则从配置文件读取默认值
        if auto is None:
            auto = DATA_PATHS.get('auto_generate_xpath', True)

        # 如果传入的是字符串且不是定位器元组，则从 positioning 中获取
        if isinstance(key, str) and not key.startswith(('//', '/', 'xpath=', '(')):
            selector_value = self.positioning.get(key)
            if selector_value is None:
                raise KeyError(f"元素 key '{key}' 不存在")
            key_name = key  # 记录使用的 key 名称
        else:
            selector_value = key
            key_name = None

        selectors = selector_value if isinstance(selector_value, list) else [selector_value]

        # 遍历每个 selector 进行尝试
        for selector_item in selectors:
            # 跳过空字符串
            if not selector_item or (isinstance(selector_item, str) and not selector_item.strip()):
                continue

            selector_item_locator = self._identify_locator(selector_item)
            try:
                # 如果指定了 index，使用 XPath 索引（带括号格式）
                if index is not None and isinstance(selector_item_locator, tuple) and selector_item_locator[0] == By.XPATH:
                    # 检查 XPath 是否已经包含索引
                    if not re.search(r'\[\d+\]$', selector_item_locator[1]):
                        # 添加索引（使用括号格式）
                        selector_item_locator = (By.XPATH, f"({selector_item_locator[1]})[{index}]")
                        print(f"🎯 使用 XPath 索引：({selector_item_locator[1]})[{index}]")

                elem = WebDriverWait(self.driver, DATA_PATHS['wait_time'], poll_frequency=0.5).until(
                    EC.element_to_be_clickable(selector_item_locator)
                )
                elem.clear()
                elem.send_keys(text)

                # 截图
                screenshot_desc = desc if desc else f"输入：{text} 到 {key_name if key_name else selector_item_locator}"
                if index:
                    screenshot_desc = f"输入：{text} 到 {key_name if key_name else selector_item_locator} (第{index}个)"
                self.take_step_screenshot(screenshot_desc)

                # 收集 desc 信息用于后续生成 XPath
                if desc and key_name and auto:
                    # 先检查是否已有 HTML 文件路径
                    html_path = None
                    if desc_collector:
                        last_record = desc_collector[-1]
                        if last_record.get('html_path') and os.path.exists(last_record['html_path']):
                            html_path = last_record['html_path']

                    # 如果没有有效的 HTML 文件，保存一个新的（首次获取）
                    if not html_path:
                        # print(f"📄 首次保存 HTML - URL: {self.driver.current_url}")
                        html_path = self.save_html_code(file_name='elem/auto_generated.html')

                    # 收集 desc 信息（带上 HTML 路径、index、tag 和父元素、祖父元素信息）
                    self.collect_desc_info(key_name, desc, html_path, index=index, tag=tag, p_tag=p_tag, p_name=p_name, o_tag=o_tag, o_name=o_name)

                    # 自动生成并更新 XPath（实时保存，传递 index、tag 和父元素、祖父元素信息）
                    self.auto_and_update_xpath_realtime(key=key_name, desc=desc, html_path=html_path, index=index, tag=tag, p_tag=p_tag, p_name=p_name, o_tag=o_tag, o_name=o_name)
                elif desc and key_name and not auto:
                    # 不生成定位，但记录 desc 信息
                    pass
                    # print(f"ℹ️  已跳过自动生成 - key: {key_name}, desc: {desc}")

                return  # 成功输入后直接返回
            except (TimeoutException, ElementNotVisibleException, InvalidElementStateException, InvalidSelectorException) as e:
                # 如果是非法选择器错误，打印详细信息
                if isinstance(e, InvalidSelectorException):
                    print(f"⚠️  选择器错误 - selector_item: {selector_item}, locator: {selector_item_locator}")
                continue  # 继续尝试下一个 selector

        # 如果所有 selector 都尝试失败，则抛出异常
        error_msg = f"无法在任何提供的元素中输入文本：{selector_item_locator}"
        raise RuntimeError(error_msg)

    def scroll(self, element):
        """
        滚动到指定元素
        :param element: 元素定位器 (By, value)、定位器列表、元素 key 字符串，或 WebElement 对象
        :return: 当前对象自身
        """
        self.wait_time()

        def find_with_exc():
            # 如果已经是 WebElement 对象，直接滚动
            from selenium.webdriver.remote.webelement import WebElement
            if isinstance(element, WebElement):
                self.driver.execute_script("arguments[0].scrollIntoView();", element)
                return

            # 如果传入的是字符串且不是定位器元组，则从 positioning 中获取
            if isinstance(element, str) and not element.startswith(('//', '/', 'xpath=', '(')):
                selector_value = self.positioning.get(element)
                if selector_value is None:
                    raise KeyError(f"元素 key '{element}' 不存在")
            else:
                selector_value = element

            selectors = selector_value if isinstance(selector_value, list) else [selector_value]

            for selector_item in selectors:
                # 跳过空字符串
                if not selector_item or (isinstance(selector_item, str) and not selector_item.strip()):
                    continue

                locator = self._identify_locator(selector_item)
                try:
                    wait = WebDriverWait(self.driver, DATA_PATHS['wait_time'], poll_frequency=0.5)
                    elem = wait.until(EC.presence_of_element_located(locator))
                    self.driver.execute_script("arguments[0].scrollIntoView();", elem)
                    return
                except (TimeoutException, NoSuchElementException, InvalidSelectorException) as e:
                    # 如果是非法选择器错误，打印详细信息
                    if isinstance(e, InvalidSelectorException):
                        print(f"⚠️  选择器错误 - selector_item: {selector_item}, locator: {locator}")
                    continue

            error_msg = f"无法滚动到任何提供的元素：{selectors}"
            raise RuntimeError(error_msg)

        return self.exc(find_with_exc)

    def clear(self, element):
        """
        清除输入框内容
        :param element: 定位器 (By, value)
        :return: None
        """

        def find_with_exc():
            wait = WebDriverWait(self.driver, DATA_PATHS['wait_time'], poll_frequency=0.5)
            return wait.until(EC.element_to_be_clickable(element)).clear()

        return self.exc(find_with_exc)

    def upload_file(self, key=None, file_path=None, desc=None):
        """
        上传文件
        :param key: 元素定位器 (By, value) 或定位器列表，或元素 key 字符串
        :param file_path: 文件路径
        :param desc: 注解以及截图的名称（可选，未提供时自动生成）
        :return: None
        """

        def find_with_exc():
            # 如果传入的是字符串且不是定位器元组，则从 positioning 中获取
            if isinstance(key, str) and not key.startswith(('//', '/', 'xpath=', '(')):
                selector_value = self.positioning.get(key)
                if selector_value is None:
                    raise KeyError(f"元素 key '{key}' 不存在")
                key_name = key
            else:
                selector_value = key
                key_name = None

            selectors = selector_value if isinstance(selector_value, list) else [selector_value]

            for selector_item in selectors:
                # 跳过空字符串
                if not selector_item or (isinstance(selector_item, str) and not selector_item.strip()):
                    continue

                locator = self._identify_locator(selector_item)
                try:
                    element = WebDriverWait(self.driver, DATA_PATHS['wait_time'], poll_frequency=0.5).until(
                        EC.presence_of_element_located(locator)
                    )
                    self.wait_time()
                    element.send_keys(file_path)

                    # 截图
                    screenshot_desc = desc if desc else f"上传文件：{file_path} 到 {key_name if key_name else locator}"
                    self.take_step_screenshot(screenshot_desc)

                    return
                except (TimeoutException, ElementNotVisibleException, InvalidElementStateException, InvalidSelectorException) as e:
                    # 如果是非法选择器错误，打印详细信息
                    if isinstance(e, InvalidSelectorException):
                        print(f"⚠️  选择器错误 - selector_item: {selector_item}, locator: {locator}")
                    continue

            error_msg = f"无法在任何提供的元素中上传文件：{selectors}"
            raise RuntimeError(error_msg)

        return self.exc(find_with_exc)

    def file_path(self, file_type):
        """
        获取指定类型的文件路径
        :param file_type: 文件类型
        :return: 文件路径
        """
        self.wait_time()
        return self.file_paths[file_type]

    def update_excel_with_number(self, file_type, number, row=2, column=1):
        """
        更新 Excel 文件中的字段数据
        :param file_type: 文件类型
        :param number: 新的数值
        :param row: 写入的行号，默认为 2
        :param column: 写入的列号，默认为 1
        :return: 当前对象自身
        """
        excel_file_path = self.file_path(file_type)
        workbook = load_workbook(excel_file_path)
        sheet = workbook['Sheet1']
        sheet.cell(row=row, column=column, value=number)
        workbook.save(excel_file_path)
        return self

    def update_and_modify_the_imei(self, file_type, row=2, column=3):
        """
        新增菜单单专用
        更新 Excel 文件中的字段数据
        修改 imei 号
        """
        excel_file_path = self.file_path(file_type)
        workbook = load_workbook(excel_file_path)
        sheet = workbook['Sheet1']
        sheet.cell(row=row, column=column, value=self.imei)
        workbook.save(excel_file_path)
        return self

    # -------------------------------------------------------------------------------------------------
    def get_network_logs(self, url_pattern=None):
        """
        获取 Chrome 浏览器的网络请求日志
        :param url_pattern: URL 匹配模式（可选），如果提供则只返回包含该模式的请求
        :return: 网络请求日志列表
        """
        logs = self.driver.get_log('performance')
        network_logs = []

        for log in logs:
            try:
                message = json.loads(log['message'])
                method = message['message']['method']

                # 只处理 Network.requestWillBeSent 和 Network.responseReceived 事件
                if method in ['Network.requestWillBeSent', 'Network.responseReceived']:
                    params = message['message']['params']
                    request_id = params.get('requestId')
                    request = params.get('request', {})

                    log_entry = {
                        'request_id': request_id,
                        'method': method,
                        'url': request.get('url', ''),
                        'request_method': request.get('method', ''),
                        'headers': request.get('headers', {}),
                        'post_data': request.get('postData', '')
                    }

                    # 如果有 postData，尝试解析为 JSON
                    if log_entry['post_data']:
                        try:
                            log_entry['post_data_json'] = json.loads(log_entry['post_data'])
                        except:
                            log_entry['post_data_raw'] = log_entry['post_data']

                    # 如果提供了 URL 模式，进行过滤
                    if url_pattern and url_pattern not in log_entry['url']:
                        continue

                    network_logs.append(log_entry)

            except Exception as e:
                continue

        return network_logs

    def capture_api_request(self, url_keyword, save_filename=None, exact_match=True):
        """
        捕获并验证 API 请求参数
        【修改点】：使用字典存储，相同 url_keyword 会自动覆盖旧数据，只保留最后一次请求
        :param url_keyword: URL 关键字或完整 URL
        :param save_filename: 保存的文件名，如果为 None 则自动使用调用方法名
        :param exact_match: 是否精确匹配 URL（True 时为完整 URL 匹配，False 时为关键字匹配）
        :return: 捕获到的请求数据（只返回 request_body 格式化的数据）
        """
        # 如果未指定文件名，自动获取调用者方法名
        if save_filename is None:
            try:
                # 向上查找调用栈，跳过装饰器 wrapper
                caller_frame = inspect.currentframe().f_back
                while caller_frame and caller_frame.f_code.co_name in ['wrapper', '_validate_data_with_cache']:
                    caller_frame = caller_frame.f_back
                method_name = caller_frame.f_code.co_name
                save_filename = f"{method_name}.json"
                print(f"📝 自动使用调用方法名作为文件名：{save_filename}")
            except Exception as e:
                print(f"⚠️  获取调用方法名失败：{e}，使用默认文件名")
                save_filename = 'api_capture.json'

        # 等待一下确保所有请求都完成
        self.wait_time(2)

        # 获取网络日志
        logs = self.driver.get_log('performance')
        # 【修改点】：不再使用列表追加，而是定义一个变量暂存当前 keyword 的最新一次请求
        latest_request_data = None

        for log in logs:
            try:
                message = json.loads(log['message'])
                method = message['message']['method']

                # 只处理 Network.requestWillBeSent 事件
                if method == 'Network.requestWillBeSent':
                    params = message['message']['params']
                    request = params.get('request', {})
                    url = request.get('url', '')

                    # 检查 URL 匹配
                    url_matched = False
                    if exact_match:
                        # 精确匹配完整 URL
                        url_matched = (url == url_keyword)
                    else:
                        # 关键字匹配
                        url_matched = (url_keyword in url)

                    if url_matched:
                        # 处理 POST 数据
                        post_data = request.get('postData')
                        if post_data:
                            try:
                                # 解析为 JSON 对象
                                request_body = json.loads(post_data)

                                # 构建请求数据
                                current_request_data = {
                                    'url': url,
                                    'method': request.get('method', 'GET'),
                                    'timestamp': log.get('timestamp', ''),
                                    'request_body': request_body
                                }

                                # 【关键修改】：直接更新变量，循环结束后自然保留最后一次匹配的数据
                                latest_request_data = current_request_data

                            except json.JSONDecodeError as e:
                                print(f"⚠️  JSON 解析失败：{e}")
                                continue
                        else:
                            # 没有 POST 数据的请求不保存
                            continue

            except Exception as e:
                continue

        # 【修改点】：将最新的一次请求存入字典，覆盖旧数据
        if latest_request_data:
            self._captured_requests_map[url_keyword] = latest_request_data
            # 构造保存数据列表（兼容原有的保存逻辑，但只包含最新的一条）
            self._save_capture_to_json([latest_request_data], save_filename)
        else:
            print(f"⚠️  未捕获到匹配关键字 '{url_keyword}' 的请求")

        return [latest_request_data] if latest_request_data else []

    def _save_capture_to_json(self, data, filename):
        """
        将捕获的数据保存到 JSON 文件 - 只保存 request_body 部分
        :param data: 要保存的数据列表（现在通常只包含最新的一条）
        :param filename: 文件名
        :return: 文件路径
        """
        try:
            # 使用 cache_assert 目录
            cache_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                'cache_parameters'
            )
            os.makedirs(cache_path, exist_ok=True)

            file_path = os.path.join(cache_path, filename)

            # 提取 request_body 部分
            if data and len(data) > 0:
                # 由于 capture_api_request 已经确保只传入最新的一条数据
                # 这里直接取第一条的 request_body 即可，避免生成列表包裹
                output_data = data[0].get('request_body', {})
            else:
                output_data = {}

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=4)

            print(f"✅ API 请求参数已捕获并保存到：{file_path}")
            # 打印前 200 个字符用于调试
            print(f"📄 保存的数据格式：{json.dumps(output_data, ensure_ascii=False, indent=2)[:200]}...")
            return file_path

        except Exception as e:
            print(f"❌ 保存捕获数据失败：{e}")
            raise

    def clear_network_logs(self):
        """
        清空网络日志
        :return: None
        """
        self.driver.execute_cdp_cmd('Network.clearBrowserCache', {})
        self.driver.execute_cdp_cmd('Network.clearBrowserCookies', {})

    def enable_network_log(self):
        """
        启用网络日志记录
        :return: None
        """
        self.driver.execute_cdp_cmd('Network.enable', {})

    def disable_network_log(self):
        """
        禁用网络日志记录
        :return: None
        """
        self.driver.execute_cdp_cmd('Network.disable', {})


class ImportDataEdit(BasePage, InitializeParams):
    def __init__(self, driver):
        super().__init__(driver)
        self.file_paths = {}  # 文件路径

    def get_inventory_data(self, file_type, data_type, i=None, j=None, row=2, column=1):
        """
        库存管理 | 库存列表
        获取库存数据并更新到 Excel
        :param file_type: 文件类型
        :param data_type: 数据类型 ('imei' 或 'articles_no')
        :param i: 库存状态 2 库存中 1 待入库 3 已出库
        :param j: 物品状态 13 待销售 3 待分货 7 维修中 5 质检中 19 销售预售中 14 销售铺货中 16 待送修 9 已销售 15 销售售后中 17 送修中
                  11 采购售后完成 12 采购售后中 18 仅出库 10 待采购售后 1 待发货
        :param row: 写入的行号，默认为 2
        :param column: 写入的列号，默认为 1
        """
        if data_type not in ['imei', 'articles_no']:
            raise ValueError(f"不支持的数据类型：{data_type}")
        kwargs = {k: v for k, v in {'i': i, 'j': j}.items() if v is not None}
        res = self.pc.UYV6mZaVwDk4HHhyuWRRp(**kwargs)
        number = res[0][data_type]
        self.copy(number)
        return self.update_excel_with_number(file_type, number, row=row, column=column)

    def get_attachment_data(self, file_type, row=2, column=1):
        """
        配件管理 | 配件库存 | 库存列表
        获取库存中 物品编号
        :param file_type: 文件类型
        :param row: 写入的行号，默认为 2
        :param column: 写入的列号，默认为 1
        """
        articles_no = self.pc.CtRBRcFNn2LnUPfJF5Yhu(i='2')[0]['articlesNo']
        self.copy(articles_no)
        self.update_excel_with_number(file_type, articles_no, row=row, column=column)  # 更新 Excel
        return articles_no  # 返回物品编号


if __name__ == '__main__':
    a = ImportDataEdit("!")


def reset_after_execution(func):
    """装饰器：执行完方法后重置浏览器到主页状态"""

    @wraps(func)
    def wrapper(self, *args, **kwargs):
        try:
            result = func(self, *args, **kwargs)
            # 重置到主页
            home_url = URL['web']['index'].replace('/epbox_erp', '')  # ui 首页 url
            self.driver.get(home_url)
            self.wait_time(1)
            return result
        except Exception as e:
            # 即使出错也要重置状态
            home_url = URL['web']['index'].replace('/epbox_erp', '')  # ui 首页 url
            self.driver.get(home_url)
            raise e

    return wrapper
