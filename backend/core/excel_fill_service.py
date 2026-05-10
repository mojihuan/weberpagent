"""Excel template filling service for test automation runs.

Copies Excel templates to per-run isolated directories and fills cells
via openpyxl. Supports method chaining for sequential cell writes.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook


class ExcelFillService:
    """Manages Excel template copying and cell filling for a single run."""

    def __init__(self, templates_dir: str, filled_dir: str, run_id: str) -> None:
        self._templates_dir = Path(templates_dir)
        self._filled_dir = Path(filled_dir)
        self._run_id = run_id
        self._filled_paths: dict[str, Path] = {}

    def _template_source(self, template_name: str) -> Path:
        """Resolve the source path for a template, trying with and without .xlsx suffix."""
        candidate = self._templates_dir / template_name
        if candidate.exists():
            return candidate
        xlsx_candidate = self._templates_dir / f"{template_name}.xlsx"
        if xlsx_candidate.exists():
            return xlsx_candidate
        raise FileNotFoundError(f"Template not found: {template_name}")

    def _ensure_copy(self, template_name: str) -> Path:
        """Copy template to run directory on first access, return filled path."""
        if template_name in self._filled_paths:
            return self._filled_paths[template_name]

        source = self._template_source(template_name)
        run_dir = self._filled_dir / self._run_id
        run_dir.mkdir(parents=True, exist_ok=True)

        dest = run_dir / source.name
        shutil.copy2(source, dest)

        self._filled_paths[template_name] = dest
        return dest

    def fill_excel(
        self,
        template_name: str,
        sheet: str = "Sheet1",
        row: int = 2,
        col: int = 1,
        value: object = None,
    ) -> ExcelFillService:
        """Fill a cell in an Excel template copy.

        Copies the template on first call for this template_name.
        Returns self for method chaining.
        """
        dest = self._ensure_copy(template_name)

        wb = load_workbook(dest)
        ws = wb[sheet]
        ws.cell(row=row, column=col, value=value)
        wb.save(dest)
        wb.close()

        return self

    def get_excel_path(self, template_name: str) -> Path:
        """Return absolute path to the filled copy, or original template if not yet filled."""
        if template_name in self._filled_paths:
            return self._filled_paths[template_name]
        return self._template_source(template_name)
