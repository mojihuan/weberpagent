"""Excel 模版生成器

生成标准化的 .xlsx 测试用例模版，包含列头、示例数据、README 说明页和数据验证。

TEMPLATE_COLUMNS 是模版生成器和解析器的共享列合约（单一数据源）。
"""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_COLUMNS = [
    {"key": "name", "header": "任务名称", "width": 25, "required": True, "default": None, "aliases": []},
    {"key": "login_role", "header": "登录角色", "width": 15, "required": False, "default": None, "aliases": []},
    {"key": "description", "header": "任务描述", "width": 40, "required": True, "default": None, "aliases": ["测试步骤"]},
    {"key": "target_url", "header": "目标URL", "width": 35, "required": False, "default": "", "aliases": []},
    {"key": "max_steps", "header": "最大步数", "width": 12, "required": False, "default": 10, "aliases": []},
    {"key": "preconditions", "header": "前置条件", "width": 40, "required": False, "default": None, "aliases": []},
    {"key": "assertions", "header": "断言", "width": 50, "required": False, "default": None, "aliases": []},
]

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_EXAMPLE_ROW_FULL = [
    "登录功能测试",
    "main",
    "打开登录页面，输入用户名和密码，点击登录按钮，验证是否跳转到首页",
    "https://erp.example.com/login",
    15,
    '["context[\'token\'] = login_api()"]',
    '[{"methodName":"check_login_status","headers":"main"}]',
]

_EXAMPLE_ROW_MINIMAL = [
    "创建订单测试",
    None,
    "登录后进入订单页面，填写订单信息并提交，验证订单创建成功",
    "https://erp.example.com/orders/new",
    20,
    None,
    None,
]

_README_CONTENT = [
    (1, "测试用例 Excel 模版说明"),
    (3, "一、列说明"),
    (4, "列名 | 类型 | 是否必填 | 说明"),
    (5, "任务名称 | 文本 | 必填 | 测试用例的名称，1-200 个字符"),
    (6, "登录角色 | 文本 | 选填 | ERP 登录角色（如 main），留空则手动登录"),
    (7, "任务描述 | 文本 | 必填 | 测试步骤的自然语言描述"),
    (8, "目标URL | 文本 | 选填 | 测试起始页面地址，留空使用系统默认地址"),
    (9, "最大步数 | 整数 | 选填 | 最大执行步数（1-100），默认 10"),
    (10, "前置条件 | JSON 数组 | 选填 | Python 代码列表，如 [\"code1\", \"code2\"]"),
    (11, "断言 | JSON 数组 | 选填 | 断言配置列表，如 [{\"methodName\":\"xxx\",\"headers\":\"main\"}]"),
    (14, "二、前置条件格式"),
    (15, '前置条件使用 JSON 数组格式，填写 Python 代码字符串。例如：["context[\'token\'] = login_api()"]'),
    (17, "三、断言格式"),
    (18, '断言使用 JSON 数组格式，每个元素是一个对象。例如：[{"methodName":"check_login_status","headers":"main"}]'),
    (20, "四、注意事项"),
    (21, "1. 任务名称和任务描述为必填项"),
    (22, "2. 最大步数范围 1-100，默认 10"),
    (23, "3. 前置条件和断言为可选项，留空即可"),
    (24, "4. 请勿合并单元格"),
    (25, "5. 填写完成后保存文件并上传"),
]


def generate_template() -> BytesIO:
    """生成测试用例 Excel 模版并返回 BytesIO 缓冲区。

    Returns:
        BytesIO: 包含完整 .xlsx 模版的内存缓冲区，已 seek 到起始位置。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    _write_styled_headers(ws)
    _set_column_widths(ws)
    _add_max_steps_validation(ws)
    _write_example_rows(ws)
    ws.freeze_panes = "A2"
    _create_readme_sheet(wb)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_styled_headers(ws: Any) -> None:
    """写入带样式的列头行（第一行）。"""
    for idx, col_def in enumerate(TEMPLATE_COLUMNS):
        cell = ws.cell(row=1, column=idx + 1, value=col_def["header"])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _set_column_widths(ws: Any) -> None:
    """根据 TEMPLATE_COLUMNS 设置列宽。"""
    for idx, col_def in enumerate(TEMPLATE_COLUMNS):
        col_letter = get_column_letter(idx + 1)
        ws.column_dimensions[col_letter].width = col_def["width"]


def _add_max_steps_validation(ws: Any) -> None:
    """为 max_steps 列（D 列）添加数据验证。"""
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=100,
        allow_blank=True,
    )
    dv.error = "请输入 1-100 之间的整数"
    dv.errorTitle = "输入错误"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    dv.prompt = "请输入 1 到 100 之间的整数"
    dv.promptTitle = "最大步数"
    ws.add_data_validation(dv)
    dv.add("E2:E10000")


def _write_example_rows(ws: Any) -> None:
    """写入 2 行示例数据。"""
    for col_idx, value in enumerate(_EXAMPLE_ROW_FULL):
        ws.cell(row=2, column=col_idx + 1, value=value)

    for col_idx, value in enumerate(_EXAMPLE_ROW_MINIMAL):
        ws.cell(row=3, column=col_idx + 1, value=value)


def _create_readme_sheet(wb: Any) -> None:
    """创建 README 说明页。"""
    ws_readme = wb.create_sheet("README")

    for row_num, content in _README_CONTENT:
        cell = ws_readme.cell(row=row_num, column=1, value=content)
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
