"""Excel file parser for test case import.

Parses .xlsx files into structured ParsedRow objects with collect-all error
handling. Uses TEMPLATE_COLUMNS from excel_template as column contract.
"""

from dataclasses import dataclass, field
from typing import Any

from io import BytesIO
import json

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from backend.utils.excel_template import TEMPLATE_COLUMNS


@dataclass(frozen=True)
class ParsedRow:
    """A single parsed row with data and any errors."""

    row_number: int
    data: dict[str, Any]
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class ParseResult:
    """Result of parsing an Excel workbook."""

    rows: list[ParsedRow]
    total_rows: int
    has_errors: bool


def _coerce_string(value: Any) -> str | None:
    """Coerce cell value to string. Returns None for empty cells."""
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value).lower()
    return str(value).strip()


def _coerce_int(value: Any, default: int | None = None) -> int | None:
    """Coerce cell value to int. Returns default for empty cells."""
    if value is None:
        return default
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return default
        try:
            return int(float(stripped))
        except (ValueError, TypeError):
            return None
    return None


def _coerce_json_list(value: Any) -> tuple[list | None, str | None]:
    """Parse JSON array from cell. Returns (parsed_list_or_None, error_or_None)."""
    if value is None:
        return None, None
    raw = str(value).strip()
    if not raw:
        return None, None
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed, None
        return None, f"JSON 值不是数组: {raw[:50]}"
    except json.JSONDecodeError:
        return None, f"JSON 格式错误: {raw[:50]}"


def _is_empty_row(cells: tuple) -> bool:
    """Check if all cells in a row are None (never entered by user).

    Rows where the user typed something (even an empty string) are NOT
    considered empty -- they should be parsed so required-field validation
    can report errors.
    """
    for cell in cells:
        if cell is None:
            continue
        if isinstance(cell, MergedCell):
            # MergedCell means the row has content (user merged cells)
            return False
        if cell.value is not None:
            return False
    return True


def _has_merged_cells(cells: tuple) -> bool:
    """Check if any cell in the row is a MergedCell."""
    return any(isinstance(cell, MergedCell) for cell in cells)


def _detect_old_template(ws: Any) -> bool:
    """Detect if the workbook uses old 6-column format (no login_role).

    Old template headers: 任务名称, 任务描述, 目标URL, 最大步数, 前置条件, 断言
    New template headers: 任务名称, 登录角色, 任务描述, ...
    The key difference: column 2 is 任务描述 (old) vs 登录角色 (new).
    """
    col2 = ws.cell(row=1, column=2).value
    if col2 and str(col2).strip() == "任务描述":
        return True
    return False


def _validate_headers(ws: Any) -> list[str] | None:
    """Validate that row 1 headers match TEMPLATE_COLUMNS.

    Lenient: allows FEWER columns than TEMPLATE_COLUMNS (old templates).
    Also supports old 6-column templates (no login_role) via _detect_old_template.
    Supports column aliases (e.g. '测试步骤' as alias for '任务描述').
    Returns list of error strings if mismatch, None if headers are valid.
    """
    errors = []
    actual_col_count = ws.max_column or 0

    # Old template: skip login_role column in validation
    if _detect_old_template(ws):
        check_columns = [col for col in TEMPLATE_COLUMNS if col["key"] != "login_role"]
    else:
        check_columns = list(TEMPLATE_COLUMNS)

    check_count = min(actual_col_count, len(check_columns))

    for idx in range(check_count):
        expected_header = check_columns[idx]["header"]
        aliases = check_columns[idx].get("aliases", [])
        actual = ws.cell(row=1, column=idx + 1).value
        actual_str = str(actual).strip() if actual else ""

        if actual_str and (actual_str == expected_header or actual_str in aliases):
            continue

        errors.append(
            f"列 {idx + 1} 表头应为 '{expected_header}'，实际为 '{actual}'"
        )

    extra_cols = actual_col_count - len(check_columns)
    if extra_cols > 0:
        errors.append(f"多余列: 发现 {actual_col_count} 列，预期 {len(check_columns)} 列")

    return errors if errors else None


def _extract_cell_value(row_tuple: tuple, data_col_idx: int) -> Any:
    """Extract cell value from row tuple, handling MergedCell."""
    if data_col_idx >= len(row_tuple):
        return None
    cell = row_tuple[data_col_idx]
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def _coerce_column_value(
    key: str, cell_value: Any, col_def: dict, row_number: int,
) -> tuple[Any, str | None]:
    """Coerce a single column value. Returns (value, error_or_None)."""
    if key == "login_role":
        return _coerce_string(cell_value), None
    if key in ("name", "description", "target_url"):
        coerced = _coerce_string(cell_value)
        if key == "target_url" and coerced is None:
            return col_def.get("default", ""), None
        return coerced if coerced is not None else (col_def.get("default") or None), None
    if key == "max_steps":
        coerced = _coerce_int(cell_value, default=col_def.get("default", 10))
        if coerced is None and cell_value is not None and str(cell_value).strip() != "":
            return col_def.get("default", 10), "最大步数必须为 1-100 之间的整数"
        return coerced, None
    if key in ("preconditions", "assertions"):
        parsed_list, json_error = _coerce_json_list(cell_value)
        if json_error is not None:
            fallback = str(cell_value).strip() if cell_value is not None else None
            return fallback, f"{col_def['header']}: {json_error}"
        return parsed_list, None
    return cell_value, None


def _parse_single_row(
    row_tuple: tuple,
    col_mapping: list[tuple[int, dict]],
    is_old_template: bool,
    row_number: int,
) -> ParsedRow:
    """Parse a single Excel row into a ParsedRow with data and errors."""
    row_errors: list[str] = []

    if _has_merged_cells(row_tuple):
        row_errors.append("第 {} 行包含合并单元格，请取消合并后重新上传".format(row_number))

    data: dict[str, Any] = {}
    if is_old_template:
        data["login_role"] = None

    for data_col_idx, (template_col_idx, col_def) in enumerate(col_mapping):
        key = col_def["key"]
        cell_value = _extract_cell_value(row_tuple, data_col_idx)
        value, error = _coerce_column_value(key, cell_value, col_def, row_number)
        data[key] = value
        if error:
            row_errors.append(error)

    for col_def in TEMPLATE_COLUMNS:
        if col_def["required"]:
            val = data.get(col_def["key"])
            if val is None or (isinstance(val, str) and val.strip() == ""):
                row_errors.append(f"必填字段 '{col_def['header']}' 不能为空")

    return ParsedRow(row_number=row_number, data=data, errors=row_errors)


def parse_excel(buffer: BytesIO) -> ParseResult:
    """Parse an .xlsx workbook into structured row data.

    Uses collect-all strategy: collects all errors across all rows without
    raising on individual row errors. Only fatal errors (unopenable file)
    raise exceptions.

    Args:
        buffer: BytesIO containing .xlsx file data.

    Returns:
        ParseResult with all parsed rows, total count, and error flag.
    """
    wb = load_workbook(buffer, data_only=True)
    ws = wb.active

    header_errors = _validate_headers(ws)
    if header_errors is not None:
        return ParseResult(
            rows=[ParsedRow(row_number=0, data={}, errors=header_errors)],
            total_rows=1, has_errors=True,
        )

    parsed_rows: list[ParsedRow] = []
    is_old_template = _detect_old_template(ws)

    if is_old_template:
        col_mapping = [(idx, col) for idx, col in enumerate(TEMPLATE_COLUMNS) if col["key"] != "login_role"]
    else:
        col_mapping = list(enumerate(TEMPLATE_COLUMNS))

    for row_tuple in ws.iter_rows(min_row=2):
        if _is_empty_row(row_tuple):
            continue
        row_number = row_tuple[0].row if row_tuple else 0
        parsed_rows.append(_parse_single_row(row_tuple, col_mapping, is_old_template, row_number))

    has_errors = any(row.errors for row in parsed_rows)
    return ParseResult(rows=parsed_rows, total_rows=len(parsed_rows), has_errors=has_errors)
