"""Tests for ExcelFillService."""
import pytest
from pathlib import Path
from openpyxl import Workbook

from backend.core.excel_fill_service import ExcelFillService


@pytest.fixture
def template_dir(tmp_path):
    templates = tmp_path / "templates"
    templates.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="IMEI")
    ws.cell(row=1, column=2, value="物品编号")
    wb.save(templates / "test_template.xlsx")
    return templates


@pytest.fixture
def output_dir(tmp_path):
    out = tmp_path / "filled"
    out.mkdir()
    return out


@pytest.fixture
def service(template_dir, output_dir):
    return ExcelFillService(
        templates_dir=str(template_dir),
        filled_dir=str(output_dir),
        run_id="test_run_001",
    )


class TestFillExcel:
    def test_fill_single_cell(self, service, template_dir, output_dir):
        result = service.fill_excel("test_template", row=2, col=1, value="IMEI123456")
        assert result is service  # chaining

        path = service.get_excel_path("test_template")
        assert path.exists()

        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["Sheet1"]
        assert ws.cell(row=2, column=1).value == "IMEI123456"
        # Original unchanged
        orig = load_workbook(template_dir / "test_template.xlsx")
        assert orig["Sheet1"].cell(row=2, column=1).value is None

    def test_fill_multiple_cells_same_template(self, service):
        service.fill_excel("test_template", row=2, col=1, value="AAA")
        service.fill_excel("test_template", row=2, col=2, value="BBB")

        from openpyxl import load_workbook
        wb = load_workbook(service.get_excel_path("test_template"))
        ws = wb["Sheet1"]
        assert ws.cell(row=2, column=1).value == "AAA"
        assert ws.cell(row=2, column=2).value == "BBB"

    def test_fill_nonexistent_template_raises(self, service):
        with pytest.raises(FileNotFoundError, match="not_found"):
            service.fill_excel("not_found", row=1, col=1, value="x")

    def test_get_path_before_fill_returns_original(self, service, template_dir):
        path = service.get_excel_path("test_template")
        assert path == template_dir / "test_template.xlsx"

    def test_fill_creates_run_subdirectory(self, service, output_dir):
        service.fill_excel("test_template", row=1, col=1, value="x")
        expected_dir = output_dir / "test_run_001"
        assert expected_dir.exists()
        assert (expected_dir / "test_template.xlsx").exists()

    def test_fill_with_sheet_name(self, service, template_dir):
        from openpyxl import load_workbook
        wb = load_workbook(template_dir / "test_template.xlsx")
        wb.create_sheet("CustomSheet")
        wb.save(template_dir / "test_template.xlsx")

        service.fill_excel("test_template", sheet="CustomSheet", row=1, col=1, value="v")
        wb2 = load_workbook(service.get_excel_path("test_template"))
        assert wb2["CustomSheet"].cell(row=1, column=1).value == "v"

    def test_different_runs_isolated(self, template_dir, output_dir):
        svc1 = ExcelFillService(str(template_dir), str(output_dir), "run_A")
        svc2 = ExcelFillService(str(template_dir), str(output_dir), "run_B")

        svc1.fill_excel("test_template", row=2, col=1, value="RUN_A")
        svc2.fill_excel("test_template", row=2, col=1, value="RUN_B")

        from openpyxl import load_workbook
        wb1 = load_workbook(svc1.get_excel_path("test_template"))
        wb2 = load_workbook(svc2.get_excel_path("test_template"))
        assert wb1["Sheet1"].cell(row=2, column=1).value == "RUN_A"
        assert wb2["Sheet1"].cell(row=2, column=1).value == "RUN_B"
