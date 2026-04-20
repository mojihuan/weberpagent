"""Excel 模版生成器单元测试"""

import json

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from backend.utils.excel_template import TEMPLATE_COLUMNS, generate_template


class TestTemplateColumns:
    """TEMPLATE_COLUMNS constant tests."""

    def test_template_columns_has_7_entries(self):
        assert len(TEMPLATE_COLUMNS) == 7

    def test_template_columns_has_correct_keys(self):
        expected_keys = ["name", "login_role", "description", "target_url", "max_steps", "preconditions", "assertions"]
        actual_keys = [col["key"] for col in TEMPLATE_COLUMNS]
        assert actual_keys == expected_keys

    def test_template_columns_has_headers(self):
        for col in TEMPLATE_COLUMNS:
            assert "header" in col
            assert isinstance(col["header"], str)
            assert len(col["header"]) > 0

    def test_template_columns_has_widths(self):
        for col in TEMPLATE_COLUMNS:
            assert "width" in col
            assert isinstance(col["width"], int)
            assert col["width"] > 0

    def test_template_columns_required_flags(self):
        required_keys = {col["key"] for col in TEMPLATE_COLUMNS if col.get("required")}
        assert required_keys == {"name", "description"}


class TestTemplateGeneration:
    """generate_template() output tests."""

    @pytest.fixture
    def workbook(self):
        """Generate template and load as workbook."""
        buffer = generate_template()
        return load_workbook(buffer)

    @pytest.fixture
    def sheet(self, workbook):
        """Get the active (测试用例) sheet."""
        return workbook.active

    def test_template_returns_bytesio(self):
        buffer = generate_template()
        assert buffer is not None
        assert buffer.seek(0, 2) > 0  # non-zero length
        buffer.seek(0)

    def test_template_valid_xlsx(self, workbook):
        """Buffer can be opened with load_workbook without error."""
        assert workbook is not None
        assert len(workbook.sheetnames) >= 1

    def test_template_has_correct_sheet_name(self, sheet):
        assert sheet.title == "测试用例"

    def test_template_has_readme_sheet(self, workbook):
        assert "README" in workbook.sheetnames

    def test_template_headers(self, sheet):
        headers = [cell.value for cell in sheet[1]]
        expected_headers = [col["header"] for col in TEMPLATE_COLUMNS]
        assert headers == expected_headers

    def test_template_header_styling(self, sheet):
        for cell in sheet[1]:
            assert cell.font.bold is True
            assert cell.font.color.rgb == "00FFFFFF"
            assert cell.fill.start_color.rgb == "003B82F6"
            assert cell.fill.fill_type == "solid"

    def test_template_column_widths(self, sheet):
        from openpyxl.utils import get_column_letter

        for idx, col_def in enumerate(TEMPLATE_COLUMNS):
            col_letter = get_column_letter(idx + 1)
            actual_width = sheet.column_dimensions[col_letter].width
            assert actual_width == col_def["width"], (
                f"Column {col_letter} ({col_def['header']}): "
                f"expected width {col_def['width']}, got {actual_width}"
            )

    def test_template_example_row_2_full_data(self, sheet):
        row = [cell.value for cell in sheet[2]]
        assert row[0] == "登录功能测试"
        assert row[1] == "main"
        assert row[2] == "打开登录页面，输入用户名和密码，点击登录按钮，验证是否跳转到首页"
        assert row[3] == "https://erp.example.com/login"
        assert row[4] == 15

        # preconditions should be valid JSON array
        preconditions = json.loads(row[5])
        assert isinstance(preconditions, list)
        assert len(preconditions) > 0

        # assertions should be valid JSON array of objects
        assertions = json.loads(row[6])
        assert isinstance(assertions, list)
        assert len(assertions) > 0
        assert isinstance(assertions[0], dict)

    def test_template_example_row_3_minimal(self, sheet):
        row = [cell.value for cell in sheet[3]]
        assert row[0] == "创建订单测试"
        assert row[1] is None  # login_role is optional
        assert row[2] == "登录后进入订单页面，填写订单信息并提交，验证订单创建成功"
        assert row[3] == "https://erp.example.com/orders/new"
        assert row[4] == 20
        # preconditions and assertions should be empty
        assert row[5] is None or row[5] == ""
        assert row[6] is None or row[6] == ""

    def test_template_max_steps_validation(self, sheet):
        data_validations = sheet.data_validations.dataValidation
        assert len(data_validations) >= 1
        dv = data_validations[0]
        assert dv.type == "whole"
        assert dv.operator == "between"
        assert dv.formula1 == "1"
        assert dv.formula2 == "100"

    def test_template_max_steps_validation_error_message(self, sheet):
        data_validations = sheet.data_validations.dataValidation
        dv = data_validations[0]
        assert "请输入 1-100 之间的整数" in dv.error

    def test_template_max_steps_validation_range(self, sheet):
        data_validations = sheet.data_validations.dataValidation
        dv = data_validations[0]
        # sqref is a MultiCellRange; str() gives "D2:D10000"
        range_str = str(dv.sqref)
        assert "E2" in range_str
        assert "E10000" in range_str

    def test_template_freeze_panes(self, sheet):
        assert sheet.freeze_panes == "A2"

    def test_template_readme_content(self, workbook):
        readme = workbook["README"]
        # Check key sections exist
        all_values = []
        for row in readme.iter_rows():
            for cell in row:
                if cell.value is not None:
                    all_values.append(cell.value)

        all_text = " ".join(str(v) for v in all_values)
        assert "测试用例 Excel 模版说明" in all_text
        assert "列说明" in all_text
        assert "前置条件格式" in all_text
        assert "断言格式" in all_text
        assert "注意事项" in all_text
