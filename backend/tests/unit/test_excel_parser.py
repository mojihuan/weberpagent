"""ExcelParser unit tests.

Tests parse_excel() with type coercion, error collection, empty row skipping,
merged cell detection, required field validation, and template round-trip.
"""

import json
from io import BytesIO

import pytest
from openpyxl import Workbook

from backend.utils.excel_template import TEMPLATE_COLUMNS, generate_template
from backend.utils.excel_parser import ParsedRow, ParseResult, parse_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(
    headers: list[str] | None = None,
    rows: list[list] | None = None,
) -> BytesIO:
    """Build a minimal .xlsx workbook for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    if headers is None:
        headers = [col["header"] for col in TEMPLATE_COLUMNS]
    ws.append(headers)
    if rows:
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _default_headers() -> list[str]:
    return [col["header"] for col in TEMPLATE_COLUMNS]


# ---------------------------------------------------------------------------
# Task 1 core tests (TDD RED phase)
# ---------------------------------------------------------------------------


class TestParseExcelBasic:
    """Basic parsing: valid data rows, column keys, skip empty rows."""

    def test_valid_two_rows_returns_two_parsed_rows(self):
        buf = _make_workbook(rows=[
            ["Task A", None, "Description A", "https://a.com", 10, None, None],
            ["Task B", None, "Description B", "https://b.com", 20, None, None],
        ])
        result = parse_excel(buf)
        assert len(result.rows) == 2
        assert result.has_errors is False

    def test_parsed_row_data_keys_match_template_columns(self):
        buf = _make_workbook(rows=[
            ["Task A", None, "Description A", "", 10, None, None],
        ])
        result = parse_excel(buf)
        row = result.rows[0]
        expected_keys = {col["key"] for col in TEMPLATE_COLUMNS}
        assert set(row.data.keys()) == expected_keys

    def test_empty_rows_skipped(self):
        buf = _make_workbook(rows=[
            ["Task A", None, "Description A", "", 10, None, None],
            [None, None, None, None, None, None, None],  # completely empty
            ["Task C", None, "Description C", "", 15, None, None],
        ])
        result = parse_excel(buf)
        assert len(result.rows) == 2
        assert result.rows[0].data["name"] == "Task A"
        assert result.rows[1].data["name"] == "Task C"


class TestTypeCoercion:
    """Type coercion: strings, ints, defaults, floats."""

    def test_string_columns_coerced_from_number(self):
        """String columns (name, description, target_url) return strings."""
        buf = _make_workbook(rows=[
            [12345, None, "Description", "https://a.com", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["name"] == "12345"
        assert isinstance(result.rows[0].data["name"], str)

    def test_max_steps_integer_cell_returns_int(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 15, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["max_steps"] == 15
        assert isinstance(result.rows[0].data["max_steps"], int)

    def test_max_steps_empty_returns_default_10(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", None, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["max_steps"] == 10

    def test_max_steps_float_returns_int(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10.0, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["max_steps"] == 10
        assert isinstance(result.rows[0].data["max_steps"], int)


class TestJsonParsing:
    """JSON column parsing: preconditions and assertions."""

    def test_preconditions_valid_json_array(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, '["code1", "code2"]', None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["preconditions"] == ["code1", "code2"]
        assert result.rows[0].errors == []

    def test_preconditions_empty_returns_none(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["preconditions"] is None
        assert result.rows[0].errors == []

    def test_preconditions_invalid_json_reports_error(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, "not json at all", None],
        ])
        result = parse_excel(buf)
        assert len(result.rows[0].errors) == 1
        assert "JSON" in result.rows[0].errors[0]

    def test_assertions_valid_json_array(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, None, '[{"methodName":"xxx","headers":"main"}]'],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["assertions"] == [{"methodName": "xxx", "headers": "main"}]
        assert result.rows[0].errors == []

    def test_assertions_non_array_json_reports_error(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, None, '{"key": "val"}'],
        ])
        result = parse_excel(buf)
        assert len(result.rows[0].errors) == 1
        assert "数组" in result.rows[0].errors[0]


class TestErrorCollection:
    """Error collection: required fields, merged cells, collect-all."""

    def test_merged_cell_detected(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        ws.append(_default_headers())
        ws.append(["Task", None, "Desc", "", 10, None, None])
        # Row 3 has a merged cell
        ws.merge_cells("A3:B3")
        ws["A3"] = "Merged task"
        ws["C3"] = ""
        ws["D3"] = 10
        ws["E3"] = None
        ws["F3"] = None
        ws["G3"] = None
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_excel(buf)
        assert len(result.rows) == 2
        merged_row = result.rows[1]
        assert any("合并单元格" in err for err in merged_row.errors)

    def test_missing_required_field_reports_error(self):
        """Row with only target_url filled (no name/description) reports required field error."""
        buf = _make_workbook(rows=[
            [None, None, None, "https://a.com", 10, None, None],  # name and description missing
        ])
        result = parse_excel(buf)
        assert len(result.rows[0].errors) >= 1
        assert any("任务名称" in err for err in result.rows[0].errors)

    def test_multiple_errors_collected_per_row(self):
        """Row with only target_url filled reports multiple required field errors."""
        buf = _make_workbook(rows=[
            [None, None, None, "https://a.com", None, None, None],  # name and description empty, max_steps invalid
        ])
        result = parse_excel(buf)
        errors = result.rows[0].errors
        assert len(errors) >= 2
        assert any("任务名称" in err for err in errors)
        assert any("任务描述" in err for err in errors)


class TestRoundTrip:
    """Template round-trip: generate_template -> parse_excel -> zero errors."""

    def test_roundtrip_template_zero_errors(self):
        buf = generate_template()
        result = parse_excel(buf)
        assert len(result.rows) == 2
        assert result.has_errors is False
        for row in result.rows:
            assert row.errors == []

    def test_roundtrip_row1_data_matches_full_example(self):
        """First example row: all 7 columns filled with valid data."""
        buf = generate_template()
        result = parse_excel(buf)
        row = result.rows[0]
        assert row.data["name"] == "登录功能测试"
        assert row.data["login_role"] == "main"
        assert row.data["description"] == "打开登录页面，输入用户名和密码，点击登录按钮，验证是否跳转到首页"
        assert row.data["target_url"] == "https://erp.example.com/login"
        assert row.data["max_steps"] == 15
        assert isinstance(row.data["preconditions"], list)
        assert isinstance(row.data["assertions"], list)

    def test_roundtrip_row2_data_matches_minimal_example(self):
        """Second example row: required fields + target_url + max_steps only."""
        buf = generate_template()
        result = parse_excel(buf)
        row = result.rows[1]
        assert row.data["name"] == "创建订单测试"
        assert row.data["login_role"] is None
        assert row.data["description"] == "登录后进入订单页面，填写订单信息并提交，验证订单创建成功"
        assert row.data["target_url"] == "https://erp.example.com/orders/new"
        assert row.data["max_steps"] == 20
        assert row.data["preconditions"] is None
        assert row.data["assertions"] is None


# ---------------------------------------------------------------------------
# Task 2 comprehensive edge case tests
# ---------------------------------------------------------------------------


class TestHeaderValidation:
    """Header mismatch detection."""

    def test_header_mismatch_reports_error(self):
        buf = _make_workbook(headers=["A", "B", "C", "D", "E", "F", "G"])
        result = parse_excel(buf)
        assert result.has_errors is True
        assert len(result.rows) == 1
        assert result.rows[0].row_number == 0
        assert len(result.rows[0].errors) >= 1

    def test_header_partial_mismatch(self):
        """First header wrong, rest correct."""
        headers = ["Wrong"] + [col["header"] for col in TEMPLATE_COLUMNS[1:]]
        buf = _make_workbook(headers=headers)
        result = parse_excel(buf)
        assert result.has_errors is True
        assert any("任务名称" in err for err in result.rows[0].errors)


class TestStringCoercion:
    """String column coercion edge cases."""

    def test_number_in_name_column_coerced_to_string(self):
        buf = _make_workbook(rows=[
            [12345, None, "Description", "https://a.com", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["name"] == "12345"
        assert isinstance(result.rows[0].data["name"], str)

    def test_boolean_in_description_coerced_to_string(self):
        """Boolean True in description -> "true"."""
        buf = _make_workbook(rows=[
            ["Task", None, True, "https://a.com", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["description"] == "true"
        assert result.rows[0].errors == []


class TestIntCoercion:
    """max_steps integer coercion edge cases."""

    def test_float_truncated_to_int(self):
        """15.7 -> 15 (truncation, not rounding)."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 15.7, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["max_steps"] == 15

    def test_string_float_coerced_to_int(self):
        """String "10.0" in max_steps -> int 10."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", "10.0", None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["max_steps"] == 10

    def test_invalid_string_in_max_steps_reports_error(self):
        """String "abc" in max_steps -> error."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", "abc", None, None],
        ])
        result = parse_excel(buf)
        assert any("最大步数" in err for err in result.rows[0].errors)

    def test_boolean_in_max_steps_reports_error(self):
        """Boolean True in max_steps -> error (boolean is not valid int)."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", True, None, None],
        ])
        result = parse_excel(buf)
        assert any("最大步数" in err for err in result.rows[0].errors)


class TestJsonEdgeCases:
    """JSON column edge cases."""

    def test_preconditions_valid_json_list(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, '["code1", "code2"]', None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["preconditions"] == ["code1", "code2"]
        assert result.rows[0].errors == []

    def test_preconditions_not_array(self):
        """Valid JSON but not an array -> error."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, '{"key": "val"}', None],
        ])
        result = parse_excel(buf)
        assert any("数组" in err for err in result.rows[0].errors)

    def test_preconditions_invalid_json(self):
        """Not valid JSON at all -> error."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, "not json at all", None],
        ])
        result = parse_excel(buf)
        assert any("JSON" in err for err in result.rows[0].errors)

    def test_assertions_valid_json_list(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, None, '[{"methodName":"xxx"}]'],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["assertions"] == [{"methodName": "xxx"}]
        assert result.rows[0].errors == []


class TestCollectAllErrors:
    """Collect-all strategy: all errors preserved across rows."""

    def test_collect_all_errors_across_rows(self):
        """3 rows: row 2 valid, row 3 has 2 errors, row 4 has 1 error."""
        buf = _make_workbook(rows=[
            ["Valid Task", None, "Valid desc", "", 10, None, None],
            [None, None, None, "https://a.com", None, None, None],  # 2+ required errors
            [None, None, "Has desc", "", 10, "bad json", None],  # 1 required + 1 json
        ])
        result = parse_excel(buf)
        assert result.total_rows == 3
        assert result.has_errors is True
        assert result.rows[0].errors == []
        assert len(result.rows[1].errors) >= 2
        assert len(result.rows[2].errors) >= 1

    def test_empty_row_between_data_rows(self):
        """Empty row between data rows is skipped."""
        buf = _make_workbook(rows=[
            ["Task A", None, "Desc A", "", 10, None, None],
            [None, None, None, None, None, None, None],
            ["Task C", None, "Desc C", "", 15, None, None],
        ])
        result = parse_excel(buf)
        assert result.total_rows == 2

    def test_optional_fields_empty_no_errors(self):
        """Row with only name + description, all optional fields empty."""
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", None, None, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].errors == []
        assert result.rows[0].data["target_url"] == ""
        assert result.rows[0].data["max_steps"] == 10
        assert result.rows[0].data["preconditions"] is None
        assert result.rows[0].data["assertions"] is None


class TestOldTemplateCompatibility:
    """Old 6-column templates (no login_role) import with login_role=NULL."""

    def test_old_6_column_template_imports_without_errors(self):
        """Old template with 6 columns (no login_role) imports successfully."""
        headers = ["任务名称", "任务描述", "目标URL", "最大步数", "前置条件", "断言"]
        buf = _make_workbook(headers=headers, rows=[
            ["Task", "Desc", "", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.has_errors is False
        assert result.rows[0].data["login_role"] is None

    def test_old_template_name_still_parsed(self):
        headers = ["任务名称", "任务描述", "目标URL", "最大步数", "前置条件", "断言"]
        buf = _make_workbook(headers=headers, rows=[
            ["My Task", "My Desc", "", 15, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["name"] == "My Task"

    def test_new_template_with_login_role(self):
        buf = _make_workbook(rows=[
            ["Task", "main", "Description", "", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["login_role"] == "main"


class TestLoginRoleCoercion:
    """login_role string coercion edge cases."""

    def test_login_role_number_coerced_to_string(self):
        buf = _make_workbook(rows=[
            ["Task", 12345, "Desc", "", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["login_role"] == "12345"

    def test_login_role_none_stays_none(self):
        buf = _make_workbook(rows=[
            ["Task", None, "Desc", "", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["login_role"] is None

    def test_login_role_strips_whitespace(self):
        buf = _make_workbook(rows=[
            ["Task", "  main  ", "Desc", "", 10, None, None],
        ])
        result = parse_excel(buf)
        assert result.rows[0].data["login_role"] == "main"
