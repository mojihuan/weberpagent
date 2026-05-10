"""End-to-end tests for Excel template filling in precondition pipeline."""

import asyncio
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook

from backend.core.precondition_service import PreconditionService
from backend.config.settings import Settings


@pytest.fixture
def template_dir(tmp_path):
    """Create template directory with sample files."""
    td = tmp_path / "templates"
    td.mkdir()
    for name in ["purchase_order", "inventory_transfer"]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value="IMEI")
        ws.cell(row=1, column=2, value="物品编号")
        wb.save(td / f"{name}.xlsx")
    return td


@pytest.fixture
def filled_dir(tmp_path):
    fd = tmp_path / "filled"
    fd.mkdir()
    return fd


@pytest.fixture
def override_settings(template_dir, filled_dir, monkeypatch):
    """Override get_settings to use tmp_path directories.

    The real get_settings is cached by @lru_cache, so monkeypatching
    environment variables alone won't work. Patch the module-level function
    so that local imports (from backend.config.settings import get_settings)
    also resolve to our fake.
    """

    def _fake_get_settings():
        return Settings(
            templates_dir=str(template_dir),
            filled_dir=str(filled_dir),
        )

    monkeypatch.setattr(
        "backend.config.settings.get_settings", _fake_get_settings
    )


class TestPreconditionExcelFill:
    def test_fill_and_get_path(self, override_settings, template_dir, filled_dir):
        """Precondition fills Excel and stores path in context."""
        svc = PreconditionService(run_id="e2e_test_001")

        code = """
context['test_value'] = 'IMEI999'
context.fill_excel('purchase_order', row=2, col=1, value='IMEI999')
context['excel_file'] = context.get_excel_path('purchase_order')
"""
        result = asyncio.run(svc.execute_single(code, 0))
        assert result.success, f"Precondition failed: {result.error}"

        ctx = svc.get_context()
        assert ctx['test_value'] == 'IMEI999'
        assert 'excel_file' in ctx
        assert 'purchase_order.xlsx' in ctx['excel_file']

        wb = load_workbook(ctx['excel_file'])
        assert wb["Sheet1"].cell(row=2, column=1).value == 'IMEI999'

    def test_variable_substitution_with_excel_path(
        self, override_settings, template_dir, filled_dir
    ):
        """Task description {{excel_file}} gets replaced with actual path."""
        svc = PreconditionService(run_id="e2e_test_002")

        code = """
context.fill_excel('purchase_order', row=2, col=1, value='ABC123')
context['excel_file'] = context.get_excel_path('purchase_order')
"""
        result = asyncio.run(svc.execute_single(code, 0))
        assert result.success

        ctx = svc.get_context()
        description = "请上传文件 {{excel_file}} 到采购入库页面"
        substituted = PreconditionService.substitute_variables(description, ctx)

        assert "{{excel_file}}" not in substituted
        assert ctx['excel_file'] in substituted
        assert substituted.startswith("请上传文件 ")

    def test_multiple_templates_in_one_task(
        self, override_settings, template_dir, filled_dir
    ):
        """One task uses multiple Excel templates."""
        svc = PreconditionService(run_id="e2e_test_003")

        code = """
context.fill_excel('purchase_order', row=2, col=1, value='A')
context['file_a'] = context.get_excel_path('purchase_order')
context.fill_excel('inventory_transfer', row=2, col=1, value='B')
context['file_b'] = context.get_excel_path('inventory_transfer')
"""
        result = asyncio.run(svc.execute_single(code, 0))
        assert result.success

        ctx = svc.get_context()
        assert 'file_a' in ctx
        assert 'file_b' in ctx
        assert ctx['file_a'] != ctx['file_b']
        assert 'purchase_order' in ctx['file_a']
        assert 'inventory_transfer' in ctx['file_b']

    def test_fill_without_run_id_raises(self):
        """Calling fill_excel without run_id raises RuntimeError."""
        svc = PreconditionService(run_id=None)

        code = "context.fill_excel('any_template', row=2, col=1, value='x')"
        result = asyncio.run(svc.execute_single(code, 0))
        assert not result.success
        assert "run_id" in result.error
